import pandas as pd
import datetime as dt
from openpyxl import load_workbook

# =========================
#  HÀM HỖ TRỢ
# =========================

def _append_or_update_raw(ws_raw, df_new):
    """
    Append dữ liệu mới vào sheet Raw.
    Nếu ngày đã có thì update, nếu chưa có thì append.
    """
    # Đọc dữ liệu cũ từ sheet Raw
    df_old = pd.DataFrame(ws_raw.values)
    df_old.columns = df_old.iloc[0]  # dòng đầu là header
    df_old = df_old.drop(0)

    # Ghép dữ liệu cũ + mới
    merged = pd.concat([df_old, df_new], ignore_index=True)

    # Ép toàn bộ cột date về datetime để tránh lỗi sort
    merged['date'] = pd.to_datetime(merged['date'], dayfirst=True, errors='coerce')

    # Sắp xếp và loại bỏ trùng lặp theo ngày (giữ bản ghi mới nhất)
    merged = merged.sort_values('date').drop_duplicates(subset=['date'], keep='last')

    # Ghi lại vào sheet Raw
    for i, col in enumerate(merged.columns, 1):
        ws_raw.cell(1, i).value = col
    for r_idx, row in enumerate(merged.values, 2):
        for c_idx, val in enumerate(row, 1):
            ws_raw.cell(r_idx, c_idx).value = val


def fetch_today_data():
    """
    Hàm giả lập fetch dữ liệu ngày hôm nay.
    Bạn có thể thay bằng API thật nếu muốn.
    """
    today = dt.date.today().strftime("%d/%m/%Y")  # dạng dd/mm/yyyy
    data = {
        "date": [today],
        "value": [42]  # ví dụ kết quả xổ số / dữ liệu nào đó
    }
    return pd.DataFrame(data)


def update_today():
    """
    Update dữ liệu hôm nay vào file Excel
    """
    # Load file Excel
    wb = load_workbook("data.xlsx")
    ws_raw = wb["Raw"]

    # Lấy dữ liệu mới
    df_new = fetch_today_data()

    # Ghi vào Raw
    _append_or_update_raw(ws_raw, df_new)

    # Lưu lại
    wb.save("data.xlsx")
    print("✅ Đã update dữ liệu hôm nay thành công!")


# =========================
#  MAIN
# =========================
if __name__ == "__main__":
    import sys
    if "--update-today" in sys.argv:
        update_today()
    else:
        print("⚠️ Thêm tham số --update-today để chạy update.")
